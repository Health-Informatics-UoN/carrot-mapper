from typing import TypedDict, List, Optional, Any, Dict, Tuple
from airflow.providers.microsoft.azure.hooks.wasb import WasbHook
from airflow.providers.amazon.aws.hooks.s3 import S3Hook
from airflow.models.connection import Connection
from airflow.utils.session import create_session
from libs.enums import StorageType
import os
from openpyxl.worksheet.worksheet import Worksheet
from libs.queries import create_fields_query
import logging
from collections import defaultdict
from airflow.providers.postgres.hooks.postgres import PostgresHook
from pathlib import Path

# PostgreSQL connection hook
pg_hook = PostgresHook(postgres_conn_id="postgres_db_conn")
# Storage type
storage_type = os.getenv("STORAGE_TYPE", StorageType.MINIO)


def get_unique_table_names(worksheet: Worksheet) -> List[str]:
    """
    Extracts unique table names from the Field Overview worksheet.

    Args:
        worksheet: The worksheet containing table names.

    Returns:
        List[str]: A list of unique table names.
    """
    # Get all the table names in the order they appear in the Field Overview page
    table_names = []
    # Iterate over cells in the first column, but because we're in ReadOnly mode we
    # can't do that in the simplest manner.
    worksheet.calculate_dimension()
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        cell_value = row[0].value
        if cell_value and isinstance(cell_value, str) and cell_value not in table_names:
            # Truncate table names because sheet names are truncated to 31 characters in Excel
            # NOTE: This can cause the table names to be duplicated
            table_names.append(cell_value[:31])
    return table_names


storage_type = os.getenv("STORAGE_TYPE", StorageType.MINIO)


def connect_to_storage():
    with create_session() as session:
        if storage_type == StorageType.AZURE:
            # Check if WASB connection exists
            existing_conn = (
                session.query(Connection)
                .filter(Connection.conn_id == "wasb_conn")
                .first()
            )

            if existing_conn is None:
                conn = Connection(
                    conn_id="wasb_conn",
                    conn_type="wasb",
                    extra={
                        "connection_string": os.getenv(
                            "AIRFLOW_VAR_WASB_CONNECTION_STRING"
                        ),
                    },
                )
                session.add(conn)
                session.commit()
                logging.info("Created new WASB connection")

        elif storage_type == StorageType.MINIO:
            # Check if MinIO connection exists
            existing_conn = (
                session.query(Connection)
                .filter(Connection.conn_id == "minio_conn")
                .first()
            )

            if existing_conn is None:
                conn = Connection(
                    conn_id="minio_conn",
                    conn_type="aws",
                    extra={
                        "endpoint_url": os.getenv("AIRFLOW_VAR_MINIO_ENDPOINT"),
                        "aws_access_key_id": os.getenv("AIRFLOW_VAR_MINIO_ACCESS_KEY"),
                        "aws_secret_access_key": os.getenv(
                            "AIRFLOW_VAR_MINIO_SECRET_KEY"
                        ),
                    },
                )
                session.add(conn)
                session.commit()
                logging.info("Created new MinIO connection")


def get_storage_hook():
    if storage_type == StorageType.AZURE:
        return WasbHook(wasb_conn_id="wasb_conn")
    elif storage_type == StorageType.MINIO:
        return S3Hook(aws_conn_id="minio_conn")
    else:
        raise ValueError(f"Unsupported storage type: {storage_type}")


def remove_BOM(intermediate: List[Dict[str, Any]]):
    """
    Given a list of dictionaries, remove any occurrences of the BOM in the keys.

    Args:
        intermediate (List[Dict[str, Any]]): List of dictionaries to remove from.

    Returns:
        The list of dictionaries with BOM removed from the keys.
    """
    return [
        {key.replace("\ufeff", ""): value for key, value in d.items()}
        for d in intermediate
    ]


def process_four_item_dict(
    four_item_data: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Dict[str, str]]]:
    """
    Converts a list of dictionaries (each with keys 'csv_file_name', 'field_name' and
    'code' and 'value') to a nested dictionary with indices 'csv_file_name',
    'field_name', 'code', and internal value 'value'.

    [{'csv_file_name': 'table1', 'field_name': 'field1', 'value': 'value1', 'code':
    'code1'},
    {'csv_file_name': 'table1', 'field_name': 'field2', 'value': 'value2', 'code':
    'code2'},
    {'csv_file_name': 'table2', 'field_name': 'field2', 'value': 'value2', 'code':
    'code2'},
    {'csv_file_name': 'table2', 'field_name': 'field2', 'value': 'value3', 'code':
    'code3'},
    {'csv_file_name': 'table3', 'field_name': 'field3', 'value': 'value3', 'code':
    'code3'}]
    ->
    {'table1': {'field1': {'value1': 'code1'}, 'field2': {'value2': 'code2'}},
    'table2': {'field2': {'value2': 'code2', 'value3': 'code3'}},
    'table3': {'field3': {'value3': 'code3'}}
    }
    """
    csv_file_names = set(row["csv_file_name"] for row in four_item_data)
    # Initialise the dictionary with the keys, and each value set to a blank dict()
    new_data_dictionary: Dict[str, Dict[str, Dict[str, str]]] = {}
    for csv_file_name in csv_file_names:
        new_data_dictionary[csv_file_name] = {}

    for row in four_item_data:
        if row["field_name"] not in new_data_dictionary[row["csv_file_name"]]:
            new_data_dictionary[row["csv_file_name"]][row["field_name"]] = {}
        new_data_dictionary[row["csv_file_name"]][row["field_name"]][row["code"]] = row[
            "value"
        ]

    return new_data_dictionary


def transform_scan_report_sheet_table(sheet: Worksheet) -> defaultdict[Any, List]:
    """
    Transforms a worksheet data into a JSON like format.
    Note: This function was copied from the workers project. More details can be found here:
    app/workers/UploadQueue/__init__.py -> function: _transform_scan_report_sheet_table

    Args:
        sheet (Worksheet): Sheet of data to transform

    Returns:
        defaultdict[Any, List]: The transformed data.
    """
    logging.debug("Start process_scan_report_sheet_table")

    sheet.calculate_dimension()
    # Get header entries (skipping every second column which is just 'Frequency')
    # So sheet_headers = ['a', 'b']
    first_row = sheet[1]
    sheet_headers = [cell.value for cell in first_row[::2]]

    d = defaultdict(list)
    for row in sheet.iter_rows(
        min_col=1,
        max_col=len(sheet_headers) * 2,
        min_row=2,
        max_row=sheet.max_row,
        values_only=True,
    ):
        # Set boolean to track whether we hit a blank row for early exit below.
        this_row_empty = True
        # Iterate across the pairs of cells in the row. If the pair is non-empty,
        # then add it to the relevant dict entry.
        for header, cell, freq in zip(sheet_headers, row[::2], row[1::2]):
            if (cell != "" and cell is not None) or (freq != "" and freq is not None):
                d[header].append((str(cell), freq))
                this_row_empty = False
        if this_row_empty:
            break
    # Clean BOM characters from keys before returning
    cleaned_dict = defaultdict(list)
    for key, value in d.items():
        clean_key = key.replace("\ufeff", "") if isinstance(key, str) else key
        cleaned_dict[clean_key] = value

    logging.debug("Finish process_scan_report_sheet_table")
    return cleaned_dict


def _default_zero(value) -> float:
    """
    Helper function that returns the input, replacing anything Falsey
    (such as Nones or empty strings) with 0.0.
    """
    return round(value or 0.0, 2)


def create_field_entries(
    worksheet: Worksheet, table_pairs: List[Tuple[str, int]]
) -> None:
    """
    Creates field entries in the database for a scan report table.

    Args:
        row: The worksheet row containing field data
        scan_report_table_id: The ID of the scan report table

    Returns:
        The ID of the newly created field
    """
    try:
        previous_row_value = None
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row + 2):
            # Guard against unnecessary rows beyond the last true row with contents
            if (previous_row_value is None or previous_row_value == "") and (
                row[0].value is None or row[0].value == ""
            ):
                break
            previous_row_value = row[0].value

            # If the row is not empty, then it is a field in a table
            if row[0].value != "" and row[0].value is not None:
                current_table_name = row[0].value
                # table_pair[0] is table name, table_pair[1] is table id
                table = next(
                    table_pair
                    for table_pair in table_pairs
                    if table_pair[0] == current_table_name
                )

                # Extract values from the row, handling possible None values
                field_name = str(row[1].value) if row[1].value is not None else ""
                description = str(row[2].value) if row[2].value is not None else ""
                type_column = str(row[3].value) if row[3].value is not None else ""
                max_length = row[4].value
                nrows = row[5].value
                nrows_checked = row[6].value

                # Calculate fractions with default values
                fraction_empty = round(_default_zero(row[7].value), 2)
                nunique_values = row[8].value
                fraction_unique = round(_default_zero(row[9].value), 2)

                # Execute the query and get the returned ID
                pg_hook.run(
                    create_fields_query,
                    parameters={
                        # table[1] is table id
                        "scan_report_table_id": table[1],
                        "name": field_name.replace("\ufeff", ""),
                        "description_column": description,
                        "type_column": type_column,
                        "max_length": max_length,
                        "nrows": nrows,
                        "nrows_checked": nrows_checked,
                        "fraction_empty": fraction_empty,
                        "nunique_values": nunique_values,
                        "fraction_unique": fraction_unique,
                    },
                )

    except Exception as e:
        logging.error(f"Error creating field entry: {str(e)}")
        raise e


def create_temp_data_dictionary_table(
    data_dictionary: Dict[str, Dict[str, Dict[str, str]]], scan_report_id: int
) -> None:
    """
    Creates a temporary table to store data dictionary information.

    Args:
        data_dictionary: A dictionary of data dictionary information
        scan_report_id: The ID of the scan report

    Returns:
        None
    """

    if not data_dictionary:
        logging.info("No data dictionary available, skipping dictionary table creation")
        return

    try:
        # Create temporary table for data dictionary
        pg_hook.run(
            """
            DROP TABLE IF EXISTS temp_data_dictionary_%(scan_report_id)s;
            CREATE TABLE temp_data_dictionary_%(scan_report_id)s (
                table_name VARCHAR(255),
                field_name VARCHAR(255),
                value TEXT,
                value_description TEXT
            )
        """,
            parameters={"scan_report_id": scan_report_id},
        )

        # Prepare data for insertion
        dictionary_records = []
        for table_name, fields in data_dictionary.items():
            for field_name, values in fields.items():
                for value, description in values.items():
                    dictionary_records.append(
                        {
                            "table_name": table_name,
                            "field_name": field_name,
                            "value": value,
                            "value_description": description,
                        }
                    )

        # Insert records into the temporary table
        if dictionary_records:
            pg_hook.insert_rows(
                table=f"temp_data_dictionary_{scan_report_id}",
                rows=[
                    (
                        d["table_name"],
                        d["field_name"],
                        d["value"],
                        d["value_description"],
                    )
                    for d in dictionary_records
                ],
                target_fields=[
                    "table_name",
                    "field_name",
                    "value",
                    "value_description",
                ],
            )

        logging.info(
            f"Created temporary data dictionary table with {len(dictionary_records)} records"
        )

    except Exception as e:
        logging.error(f"Error creating data dictionary table: {str(e)}")
        raise e


def create_temp_field_values_table(
    field_values_dict: defaultdict[Any, List], table_id: int, table_name: str
):
    """
    Creates a temporary table to store data dictionary information.

    Args:
        kwargs: Airflow context containing the data dictionary

    Returns:
        None
    """

    if not field_values_dict:
        logging.info("No field-values available, skipping field values table creation")
        return

    try:
        # Create temp table to store field value frequencies
        pg_hook.run(
            """
            DROP TABLE IF EXISTS temp_field_values_%(table_id)s;
            CREATE TABLE IF NOT EXISTS temp_field_values_%(table_id)s (
                table_name VARCHAR(255),
                field_name VARCHAR(255),
                value TEXT,
                frequency INTEGER
            )
        """,
            parameters={"table_id": table_id},
        )

        # Insert field values data into temp table
        field_values_data = []
        for field_name, values in field_values_dict.items():
            for value, frequency in values:
                # TODO: confirm about frequency of "List truncated..."
                # Convert empty strings or None to 0 for frequency
                if frequency == "" or frequency is None:
                    frequency = 0

                # Ensure frequency is an integer
                try:
                    frequency = int(frequency)
                except (ValueError, TypeError):
                    frequency = 0

                field_values_data.append(
                    {
                        "table_name": table_name,
                        "field_name": field_name,
                        "value": value,
                        "frequency": frequency,
                    }
                )

        if field_values_data:
            pg_hook.insert_rows(
                table=f"temp_field_values_{table_id}",
                rows=[
                    (
                        d["table_name"],
                        d["field_name"],
                        d["value"],
                        d["frequency"],
                    )
                    for d in field_values_data
                ],
                target_fields=[
                    "table_name",
                    "field_name",
                    "value",
                    "frequency",
                ],
            )

    except Exception as e:
        logging.error(f"Error creating data dictionary table: {str(e)}")
        raise e


# Storage hook
storage_hook = get_storage_hook()


def download_blob_to_tmp(container_name: str, blob_name: str) -> Path:
    """
    Downloads a blob from storage to a temporary local file.

    Args:
        container_name (str): The name of the storage container/bucket.
        blob_name (str): The name of the blob/file to download.

    Returns:
        Path: The local path to the downloaded file.
    """
    # TODO: double check temp file can be accessed by other tasks
    # TODO: check if temp file is persistent, if yes then we can remove the file after processing
    # https://stackoverflow.com/questions/69294934/where-is-tmp-folder-located-in-airflow
    local_path = Path(f"/tmp/{blob_name}")
    try:
        logging.info(f"Downloading file from {container_name}/{blob_name}")
        if storage_type == StorageType.AZURE:
            storage_hook.get_file(
                file_path=local_path,
                container_name=container_name,
                blob_name=blob_name,
            )
        elif storage_type == StorageType.MINIO:
            s3_object = storage_hook.get_key(key=blob_name, bucket_name=container_name)
            with open(local_path, "wb") as f:
                s3_object.download_fileobj(f)
        return local_path
    except Exception as e:
        logging.error(f"Error downloading {blob_name} from {container_name}: {str(e)}")
        raise
