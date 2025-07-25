from typing import List, Any, Dict
from openpyxl.worksheet.worksheet import Worksheet
import logging
from collections import defaultdict


def get_unique_table_names(worksheet: Worksheet) -> List[str]:
    """
    Extracts unique table names from the Field Overview worksheet.

    Args:
        worksheet: The worksheet containing table names.

    Returns:
        List[str]: A list of unique table names.
    """
    # Get all the table names in the order they appear in the Field Overview page
    table_names = []
    # Iterate over cells in the first column, but because we're in ReadOnly mode we
    # can't do that in the simplest manner.
    worksheet.reset_dimensions()  # type: ignore
    worksheet.calculate_dimension(force=True)  # type: ignore

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        cell_value = row[0].value
        if cell_value and isinstance(cell_value, str) and cell_value not in table_names:
            # Truncate table names because sheet names are truncated to 31 characters in Excel
            # NOTE: This can cause the table names to be duplicated
            table_names.append(cell_value[:31])
    return table_names


def remove_BOM(intermediate: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Given a list of dictionaries, remove any occurrences of the BOM in the keys.

    Args:
        intermediate (List[Dict[str, Any]]): List of dictionaries to remove from.

    Returns:
        The list of dictionaries with BOM removed from the keys.
    """
    return [
        {key.replace("\ufeff", ""): value for key, value in d.items()}
        for d in intermediate
    ]


def process_four_item_dict(
    four_item_data: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Dict[str, str]]]:
    """
    Converts a list of dictionaries (each with keys 'csv_file_name', 'field_name' and
    'code' and 'value') to a nested dictionary with indices 'csv_file_name',
    'field_name', 'code', and internal value 'value'.

    Note: This function was copied from the shared project. More details can be found here:
    app/shared/services/utils.py -> function: process_four_item_dict
    """
    csv_file_names = set(row["csv_file_name"] for row in four_item_data)
    # Initialise the dictionary with the keys, and each value set to a blank dict()
    new_data_dictionary: Dict[str, Dict[str, Dict[str, str]]] = {}
    for csv_file_name in csv_file_names:
        new_data_dictionary[csv_file_name] = {}

    for row in four_item_data:
        if row["field_name"] not in new_data_dictionary[row["csv_file_name"]]:
            new_data_dictionary[row["csv_file_name"]][row["field_name"]] = {}
        new_data_dictionary[row["csv_file_name"]][row["field_name"]][row["code"]] = row[
            "value"
        ]

    return new_data_dictionary


def transform_scan_report_sheet_table(sheet: Worksheet) -> defaultdict[Any, List]:
    """
    Transforms a worksheet data into a JSON like format.
    Note: This function was copied from the workers project. More details can be found here:
    app/workers/UploadQueue/__init__.py -> function: _transform_scan_report_sheet_table

    Args:
        sheet (Worksheet): Sheet of data to transform

    Returns:
        defaultdict[Any, List]: The transformed data.
    """
    logging.debug("Start process_scan_report_sheet_table")

    sheet.reset_dimensions()  # type: ignore
    sheet.calculate_dimension(force=True)  # type: ignore
    # Get header entries (skipping every second column which is just 'Frequency')
    # So sheet_headers = ['a', 'b']
    first_row = sheet[1]
    sheet_headers = [cell.value for cell in first_row[::2]]

    d = defaultdict(list)
    for row in sheet.iter_rows(
        min_col=1,
        max_col=len(sheet_headers) * 2,
        min_row=2,
        max_row=sheet.max_row,
        values_only=True,
    ):
        # Set boolean to track whether we hit a blank row for early exit below.
        this_row_empty = True
        # Iterate across the pairs of cells in the row. If the pair is non-empty,
        # then add it to the relevant dict entry.
        for header, cell, freq in zip(sheet_headers, row[::2], row[1::2]):
            if (cell != "" and cell is not None) or (freq != "" and freq is not None):
                d[header].append((str(cell), freq))
                this_row_empty = False
        if this_row_empty:
            break
    # Clean BOM characters from keys before returning
    cleaned_dict = defaultdict(list)
    for key, value in d.items():
        clean_key = key.replace("\ufeff", "") if isinstance(key, str) else key
        cleaned_dict[clean_key] = value

    logging.debug("Finish process_scan_report_sheet_table")
    return cleaned_dict


def default_zero(value) -> float:
    """
    Helper function that returns the input, replacing anything Falsey
    (such as Nones or empty strings) with 0.0.
    """
    return round(value or 0.0, 2)
