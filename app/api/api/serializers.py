import csv
from collections import Counter
from io import BytesIO, StringIO

from collections import defaultdict
from api.logger import logger

import openpyxl  # type: ignore
from datasets.serializers import DatasetSerializer
from django.contrib.auth.models import User
from drf_dynamic_fields import DynamicFieldsMixin  # type: ignore
from openpyxl.workbook.workbook import Workbook  # type: ignore
from rest_framework import serializers
from rest_framework.exceptions import NotFound, ParseError, PermissionDenied
from data.models import Concept
from mapping.models import (
    Dataset,
    ScanReport,
    ScanReportConcept,
    ScanReportField,
    ScanReportTable,
    ScanReportValue,
    VisibilityChoices,
    UploadStatus,
    MappingStatus,
)
from shared.users.serializers import UserSerializer
from mapping.permissions import has_editorship, is_admin, is_az_function_user
from shared.services.rules_export import analyse_concepts
from config.settings import DATA_UPLOAD_MAX_MEMORY_SIZE


class ConceptSerializerV2(serializers.ModelSerializer):
    class Meta:
        """
        Serializer for the Concept model.

        Serializes the concept_id, concept_name, and
        concept_code fields.

        Args:
            model (Concept): The model to be serialized.

            fields (list): The fields to be included in
            the serialized output.
        """

        model = Concept
        fields = ["concept_id", "concept_name", "concept_code"]


class UploadStatusSerializer(serializers.ModelSerializer):
    """
    Serializer for the UploadStatus model.

    Serializes the value field.
    """

    class Meta:
        model = UploadStatus
        fields = ["value"]


class MappingStatusSerializer(serializers.ModelSerializer):
    """
    Serializer for the MappingStatus model.

    Serializes the value field.

    Args:
        model (MappingStatus): The model to be serialized.

        fields (list): The fields to be included in
        the serialized output.
    """

    class Meta:
        model = MappingStatus
        fields = ["value"]


class ScanReportViewSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):
    """
    Serializer for the ScanReportViewV2, for version 2.
    Args:
        self: The instance of the class.
        data: The data to be validated.
    Returns:
        dict: The validated data for the scan report.
    Raises:
        serializers.ValidationError: If the request context is missing.
        PermissionDenied: If the user does not have the required permissions.
        NotFound: If the parent dataset is not found.
    """

    author = UserSerializer(read_only=True)
    parent_dataset = DatasetSerializer(read_only=True)
    data_partner = serializers.SerializerMethodField()
    mapping_status = MappingStatusSerializer()
    upload_status = UploadStatusSerializer(read_only=True)

    class Meta:
        model = ScanReport
        fields = (
            "id",
            "name",
            "dataset",
            "parent_dataset",
            "data_partner",
            "mapping_status",
            "upload_status",
            "upload_status_details",
            "created_at",
            "hidden",
            "author",
            "viewers",
            "editors",
            "visibility",
        )

    def get_data_partner(self, obj):
        return (
            obj.parent_dataset.data_partner.name
            if obj.parent_dataset.data_partner
            else None
        )


class ScanReportFilesSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    """
    Serializer for handling file uploads related to ScanReports.

    Validates the scan report and data dictionary files.

    Args:
        scan_report_file (FileField): The scan report file to be uploaded.
        data_dictionary_file (FileField): The data dictionary file to be uploaded.

    Raises:
        ParseError: If the file formats are incorrect or if the data dictionary
                    does not meet the required specifications.

        PermissionDenied: If the user does not have the required permissions.

        NotFound: If the parent dataset is not found.

    Returns:
        scan_report_file (FileField): The validated scan report file.
        data_dictionary_file (FileField): The validated data dictionary file.
    """

    scan_report_file = serializers.FileField(write_only=True)
    data_dictionary_file = serializers.FileField(
        write_only=True, required=False, allow_empty_file=True
    )

    class Meta:
        model = ScanReport
        fields = (
            "scan_report_file",
            "data_dictionary_file",
        )

    def validate_data_dictionary_file(self, value):
        """
        Validates the data dictionary file to ensure it is
        in CSV format and performs consistency checks.

        Args:
            value: The uploaded data dictionary file.

        Returns:
            The validated data dictionary file.
        """
        data_dictionary = value

        if str(data_dictionary) == "undefined":
            return data_dictionary

        if not str(data_dictionary).endswith(".csv"):
            raise ParseError(
                "You have attempted to upload a data dictionary "
                "which is not in CSV format. "
                "Please upload a .csv file."
            )

        # Validate Data dictionary size before attempting to upload it
        if data_dictionary.size > DATA_UPLOAD_MAX_MEMORY_SIZE:
            raise ParseError(
                f"Please upload a smaller Data dictionary. The maximum size of a Data dictionary is {DATA_UPLOAD_MAX_MEMORY_SIZE / 1024 / 1024} MB"
            )

        # Read the file once
        decoded = data_dictionary.read().decode("utf-8-sig")
        csv_reader = csv.reader(StringIO(decoded))

        errors = []

        csv_file_names = set()
        logger.info("Collecting CSV file names from the data dictionary...")

        # Check first line for correct headers to columns
        header_line = next(csv_reader)
        if header_line != ["csv_file_name", "field_name", "code", "value"]:
            raise ParseError(
                f"Dictionary file has incorrect first line. "
                f"It must be ['csv_file_name', "
                f"'field_name', 'code', 'value'], but you "
                f"supplied {header_line}. If this error is "
                f"showing extra elements, this indicates "
                f"that another line has >4 elements, "
                f"which will need to be corrected."
            )

        # Check all rows have either 3 or 4 non-empty elements, and only the 4th can be empty.
        # Start from 2 because we want to use 1-indexing _and_ skip the first row which was
        # processed above.
        for line_no, line in enumerate(csv_reader, start=2):
            line_length_nonempty = len([element for element in line if element != ""])
            if line_length_nonempty not in [3, 4]:
                errors.append(
                    ParseError(
                        f"Dictionary has "
                        f"{line_length_nonempty} "
                        f"values in line {line_no} ({line}). "
                        f"All lines must "
                        f"have either 3 or 4 entries."
                    )
                )
            # Check for whether any of the first 3 elements are empty
            for element_no, element in enumerate(line[:3], start=1):
                if element == "":
                    errors.append(
                        ParseError(
                            f"Dictionary has an empty element "
                            f"in column {element_no} in line "
                            f"{line_no}. "
                            f"Only the 4th element in any line "
                            f"may be empty."
                        )
                    )

            # Collect csv_file_name if line exists
            if line and len(line) > 0:
                csv_file_names.add(line[0].strip())

        # Checking for mismatches between csv_file_name entries and expected table names:
        logger.info("Checking for matching table names of DD and SR.")

        if not hasattr(self, "sr_table_names") or not csv_file_names:
            missing = (
                "Scan Report table names"
                if not hasattr(self, "sr_table_names")
                else "CSV file names"
            )
            logger.error(f"{missing} are missing for validation.")
            raise ValueError(f"{missing} are required for validation.")

        # Validate each csv_name against sr_table_names
        for csv_name in csv_file_names:
            if csv_name not in self.sr_table_names:
                errors.append(
                    ParseError(
                        f"'{csv_name}' in the column 'csv_file_name' in the Data Dictionary does not match any table name in the Scan Report. Hint: Make sure the extension '.csv' is included"
                        f"Available table names/CSV file names: {', '.join(sorted(self.sr_table_names))}"
                    )
                )

        # Validates the structure and content of an uploaded data dictionary CSV file.
        # This scripts checks if there are multiple records in the data dictionary with
        # the same 'csv_file_name' and 'field_name' but empty 'value'.

        if not errors:
            dd_reader = csv.DictReader(StringIO(decoded))
            field_groups = defaultdict(list)

            for record in dd_reader:
                key = (record["csv_file_name"], record["field_name"])
                field_groups[key].append(record)

            for (csv_file, field_name), records in field_groups.items():
                if len(records) > 1:
                    empty_value_count = sum(
                        1 for record in records if not record["value"]
                    )
                    if empty_value_count > 1:
                        errors.append(
                            ParseError(
                                f"Found {empty_value_count} empty 'value' mappings for "
                                f"field '{field_name}' in file '{csv_file}'. "
                                "One field in one table can only receive one vocab ID."
                            )
                        )

        if errors:
            raise ParseError(errors)

        logger.info("Data dictionary file is valid and ready for upload.")
        data_dictionary.seek(0)

        return data_dictionary

    def run_fast_consistency_checks(self, wb: Workbook):
        """
        This function performs a series of consistency checks
        on the provided Excel workbook.

        The checks are designed to quickly identify and provide
        feedback on common data issues, enabling the
        user to correct them.

        If any of these checks fail, one or a list of ParseError
        will be raised with a message detailing the issue.

        Args:
            wb (Workbook): The Excel workbook to check.

        Returns:
            True if all checks pass.

        Raises:
            ParseError: Validation checks have failed.
        """
        errors = []
        # Get the first sheet 'Field Overview'
        fo_ws = wb.worksheets[0]

        # Grab the scan report columns from the first worksheet
        # Define what the column headings should be
        source_headers = [header.value for header in fo_ws[1]]

        expected_headers = [
            "Table",
            "Field",
            "Description",
            "Type",
            "Max length",
            "N rows",
            "N rows checked",
            "Fraction empty",
            "N unique values",
            "Fraction unique",
        ]

        # Check if source headers match the expected headers. Allow unexpected
        # headers after these. This means old Scan Reports with Flag and Classification
        # columns will be handled cleanly.
        if not source_headers[:10] == expected_headers:
            errors.append(
                ParseError(
                    f"Please check the following columns exist "
                    f"in the Scan Report (Field Overview sheet) "
                    f"in this order: "
                    f"Table, Field, Description, Type, "
                    f"Max length, N rows, N rows checked, "
                    f"Fraction empty, N unique values, "
                    f"Fraction unique. "
                    f"You provided \n{source_headers[:10]}"
                )
            )
            raise ParseError(errors)

        # Check tables are correctly separated in FO - a single empty line between each
        # table
        cell_above = fo_ws["A"][1]
        for cell in fo_ws["A"][1:]:
            if (
                cell.value != cell_above.value
                and (cell.value != "" and cell.value is not None)
                and (cell_above.value != "" and cell_above.value is not None)
            ) or (cell.value == "" and cell_above.value == ""):
                errors.append(
                    ParseError(
                        f"At the cell with value {cell.value}, tables in Field Overview "
                        f"table are not correctly separated by "
                        f"a single line. "
                        f"Note: There should be no separator "
                        f"line between the header row and the "
                        f"first row of the first table."
                    )
                )
            cell_above = cell

        if errors:
            raise ParseError(errors)

        # Now that we're happy that the FO sheet is correctly formatted, we can move
        # on to comparing its contents to the sheets

        # Check tables in FO match supplied sheets
        table_names = set(
            cell.value
            for cell in fo_ws["A"][1:]
            if (cell.value != "" and cell.value is not None)
        )
        # Drop "Table Overview" and "_" sheetnames if present, as these are never used.
        table_names.difference_update(["Table Overview", "_"])

        # "Field Overview" is the only required sheet that is not a table name.
        expected_sheetnames = list(table_names) + ["Field Overview"]

        # Get names of sheet from workbook
        actual_sheetnames = set(wb.sheetnames)
        # Drop "Table Overview" and "_" sheetnames if present, as these are never used.
        actual_sheetnames.difference_update(["Table Overview", "_"])

        if sorted(actual_sheetnames) != sorted(expected_sheetnames):
            sheets_only = set(actual_sheetnames).difference(expected_sheetnames)
            fo_only = set(expected_sheetnames).difference(actual_sheetnames)
            errors.append(
                ParseError(
                    "Tables in Field Overview sheet do not "
                    "match the sheets supplied."
                )
            )
            if sheets_only:
                errors.append(
                    ParseError(
                        f"{sheets_only} are sheets that do not "
                        f"have matching entries in first column "
                        f"of the Field Overview sheet. "
                    )
                )
            if fo_only:
                errors.append(
                    ParseError(
                        f"{fo_only} are table names in first "
                        f"column of Field Overview sheet but do "
                        f"not have matching sheets supplied."
                    )
                )

        if errors:
            raise ParseError(errors)

        # Loop over the rows, and for each table, once we reach the end of the table,
        # compare the fields provided with the fields in the associated sheet
        current_table_fields = []
        current_table_name = None
        last_value = None
        for row in fo_ws.iter_rows(min_row=2):
            # Loop over rows, collecting all fields in each table in turn
            if row[0].value == "" or row[0].value is None:
                # We're at the end of the table, so process
                # Firstly, check that we're not two empty lines in a row - if so,
                # then we're beyond the last true value and iter_rows is just giving
                # us spurious rows. Abort early.
                if last_value == "" or last_value is None:
                    break
                # Get all field names from the associated sheet, by grabbing the first
                # row, and then grabbing every second column value (because the
                # alternate columns should be 'Frequency'
                table_sheet_fields = [
                    cell.value for cell in next(wb[current_table_name].rows)
                ][::2]

                # Check for multiple columns in a single sheet with the same name
                count_table_sheet_fields = Counter(table_sheet_fields)
                for field in count_table_sheet_fields:
                    if count_table_sheet_fields[field] > 1:
                        errors.append(
                            ParseError(
                                f"Sheet '{current_table_name}' "
                                f"contains more than one field "
                                f"with the name '{field}'. "
                                f"Field names must be unique "
                                f"within a table."
                            )
                        )

                # Check for multiple fields with the same name associated to a single
                # table in the Field Overview sheet
                count_current_table_fields = Counter(current_table_fields)
                for field in count_current_table_fields:
                    if count_current_table_fields[field] > 1:
                        errors.append(
                            ParseError(
                                f"Field Overview sheet contains "
                                f"more than one field with the "
                                f"name '{field}' against the "
                                f"table '{current_table_name}'. "
                                f"Field names must be unique "
                                f"within a table."
                            )
                        )

                # Check for any fields that are in only one of the Field Overview and
                # the associated sheet
                if sorted(table_sheet_fields) != sorted(current_table_fields):
                    sheet_only = set(table_sheet_fields).difference(
                        current_table_fields
                    )
                    fo_only = set(current_table_fields).difference(table_sheet_fields)
                    errors.append(
                        ParseError(
                            f"Fields in Field Overview against "
                            f"table {current_table_name} do not "
                            f"match fields in the associated "
                            f"sheet. "
                        )
                    )
                    if sheet_only:
                        errors.append(
                            ParseError(
                                f"{sheet_only} exist in the "
                                f"'{current_table_name}' sheet "
                                f"but there are no matching "
                                f"entries in the second column "
                                f"of the Field Overview sheet "
                                f"in the rows associated to the "
                                f"table '{current_table_name}'. "
                                f""
                            )
                        )
                    if fo_only:
                        errors.append(
                            ParseError(
                                f"{fo_only} exist in second "
                                f"column of Field Over"
                                f"view sheet against the table "
                                f"'{current_table_name}' but "
                                f"there are no matching column "
                                f"names in the associated sheet "
                                f"'{current_table_name}'."
                            )
                        )

                # Reset the list of fields associated to this table as we iterate down
                # the FO sheet.
                current_table_fields = []
            else:
                # Update current list of field names and the current table name - we can
                # trust the table name not to change in this case due to the earlier
                # check for empty lines between tables in the FO sheet.
                current_table_fields.append(row[1].value)
                current_table_name = row[0].value

            last_value = row[0].value

        if errors:
            raise ParseError(errors)

        return True, table_names

    def validate_scan_report_file(self, value):
        """
        Validates the scan report file to ensure it is
        in XLSX format and performs consistency checks.

        Args:
            value: The uploaded scan report file.

        Returns:
            The validated scan report file.

        Raises:
            ParseError: If the file is not in XLSX format
            or if any consistency checks fail.
        """
        scan_report = value

        if not str(scan_report).endswith(".xlsx"):
            raise ParseError(
                "You have attempted to upload a scan report which "
                "is not in XLSX format. Please upload a .xlsx file."
            )
        # Validate Scan report size before attempting to upload it
        if scan_report.size > DATA_UPLOAD_MAX_MEMORY_SIZE:
            raise ParseError(
                f"Please upload a smaller Scan report. The maximum size of a Scan report is {DATA_UPLOAD_MAX_MEMORY_SIZE / 1024 / 1024} MB"
            )

        # Load in the Excel sheet, grab the first workbook
        file_in_memory = scan_report.read()
        wb = openpyxl.load_workbook(filename=BytesIO(file_in_memory), data_only=True)

        try:
            # Store table names in the serializer
            logger.info("Collecting table names from the scan report...")
            _, self.sr_table_names = self.run_fast_consistency_checks(wb)
        except ParseError as e:
            raise e

        # If we've made it this far, the checks have passed
        return scan_report


class ScanReportCreateSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    editors = serializers.PrimaryKeyRelatedField(
        many=True, queryset=User.objects.all(), required=False
    )
    viewers = serializers.PrimaryKeyRelatedField(
        many=True, queryset=User.objects.all(), required=False
    )
    parent_dataset = serializers.PrimaryKeyRelatedField(
        queryset=Dataset.objects.order_by("name"),
        required=True,
    )
    visibility = serializers.ChoiceField(
        choices=VisibilityChoices.choices, required=True
    )

    class Meta:
        model = ScanReport
        fields = (
            "viewers",
            "editors",
            "dataset",
            "parent_dataset",
            "visibility",
        )

    def validate(self, data):
        if request := self.context.get("request"):
            if ds := data.get("parent_dataset"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(ds, request)
                    or has_editorship(ds, request)
                ):
                    raise PermissionDenied(
                        "You must be either an admin or an editor of the parent dataset to add a new scan report to it.",
                    )
            else:
                raise NotFound("Could not find parent dataset.")
        else:
            raise NotFound("Missing request context. Unable to validate scan report.")
        return super().validate(data)


class ScanReportEditSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    mapping_status = MappingStatusSerializer()

    def validate_author(self, author):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    """You must be the author of the scan report or an admin of the parent dataset
                    to change this field."""
                )
        return author

    def validate_viewers(self, viewers):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    """You must be the author of the scan report or an admin of the parent dataset
                    to change this field."""
                )
        return viewers

    def validate_editors(self, editors):
        if request := self.context.get("request"):
            if not (
                is_admin(self.instance, request) or is_az_function_user(request.user)
            ):
                raise serializers.ValidationError(
                    """You must be the author of the scan report or an admin of the parent dataset
                    to change this field."""
                )
        return editors

    def update(self, instance, validated_data):
        #  To update the "value" (not the id) of the Mapping status, the MappingStatusSerializer needs to be added,
        #  then to make changes there, the logic below is needed.
        if "mapping_status" in validated_data:
            new_mapping_status = MappingStatus.objects.get(
                value=validated_data.pop("mapping_status").pop("value")
            )
            instance.mapping_status = new_mapping_status

        return super().update(instance, validated_data)

    class Meta:
        model = ScanReport
        fields = "__all__"


class ScanReportFieldListSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):
    name = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )
    description_column = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )

    def validate(self, data):
        if request := self.context.get("request"):
            if srt := data.get("scan_report_table"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(srt, request)
                    or has_editorship(srt, request)
                ):
                    raise PermissionDenied(
                        "You must have editor or admin privileges on the scan report to edit its fields.",
                    )
            else:
                raise NotFound("Could not find the scan report table for this field.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report field."
            )
        return super().validate(data)

    class Meta:
        model = ScanReportField
        fields = "__all__"


class ScanReportTableListSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):
    date_event = ScanReportFieldListSerializerV2()
    person_id = ScanReportFieldListSerializerV2()

    class Meta:
        model = ScanReportTable
        fields = "__all__"

    def validate(self, data):
        if request := self.context.get("request"):
            if sr := data.get("scan_report"):
                if not (
                    is_az_function_user(request.user)
                    or is_admin(sr, request)
                    or has_editorship(sr, request)
                ):
                    raise PermissionDenied(
                        "You must have editor or admin privileges on the scan report to edit its tables.",
                    )
            else:
                raise NotFound("Could not find the scan report for this table.")
        else:
            raise serializers.ValidationError(
                "Missing request context. Unable to validate scan report table."
            )
        return super().validate(data)


class ScanReportTableEditSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = ScanReportTable
        fields = "__all__"


class ScanReportFieldEditSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    name = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )
    description_column = serializers.CharField(
        max_length=512, allow_blank=True, trim_whitespace=False
    )

    class Meta:
        model = ScanReportField
        fields = "__all__"


class ScanReportConceptSerializerV2(DynamicFieldsMixin, serializers.ModelSerializer):
    concept = ConceptSerializerV2()

    class Meta:
        model = ScanReportConcept
        fields = ["id", "object_id", "creation_type", "content_type", "concept"]


class ScanReportValueViewSerializerV3(serializers.ModelSerializer):
    concepts = ScanReportConceptSerializerV2(many=True, read_only=True)

    class Meta:
        model = ScanReportValue
        fields = [
            "id",
            "value",
            "frequency",
            "value_description",
            "scan_report_field",
            "concepts",
        ]


class ScanReportValueViewSerializerV2(serializers.ModelSerializer):
    class Meta:
        model = ScanReportValue
        fields = ["id", "value", "frequency", "value_description", "scan_report_field"]


class ScanReportConceptSerializer(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = ScanReportConcept
        fields = ["id", "object_id", "creation_type", "concept", "content_type"]


class GetRulesAnalysis(DynamicFieldsMixin, serializers.ModelSerializer):
    class Meta:
        model = ScanReport
        fields = "__all__"

    def to_representation(self, scan_report):
        return analyse_concepts(scan_report.id)
